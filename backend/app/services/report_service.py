"""
Serviço de Geração de Relatórios
Gera relatórios de usuários e acessos em múltiplos formatos
"""
from typing import Dict, Any, Optional, List
from datetime import datetime, timedelta
from collections import Counter, defaultdict
import csv
import io
import logging

logger = logging.getLogger(__name__)


class ReportService:
    """Serviço para geração de relatórios"""
    
    def __init__(self, db):
        self.db = db
    
    # ==================== RELATÓRIO DE USUÁRIOS ====================
    
    async def generate_users_report(
        self,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        status_filter: Optional[str] = None,  # "active", "expired", "pending", "all"
        with_image: Optional[bool] = None,
        synced_only: bool = False,
        include_cards: bool = True,
        include_access_rules: bool = True,
        format_type: str = "json"  # "json", "csv", "excel"
    ) -> Dict[str, Any]:
        """
        Gera relatório completo de usuários
        
        Args:
            start_date: Data inicial (por createdAt)
            end_date: Data final (por createdAt)
            status_filter: Filtro de status
            with_image: Filtrar por usuários com/sem imagem
            synced_only: Apenas usuários sincronizados
            include_cards: Incluir informações de cartões
            include_access_rules: Incluir regras de acesso
            format_type: Formato de saída
        
        Returns:
            Relatório formatado
        """
        logger.info("Gerando relatório de usuários...")
        start_time = datetime.now()
        
        # Construir filtros
        where = {}
        
        # Filtro por data de criação
        if start_date or end_date:
            where["createdAt"] = {}
            if start_date:
                where["createdAt"]["gte"] = start_date
            if end_date:
                where["createdAt"]["lte"] = end_date
        
        # Filtro por imagem
        if with_image is not None:
            if with_image:
                where["image"] = {"not": None}
            else:
                where["image"] = None
        
        # Filtro por sincronização
        if synced_only:
            where["idFaceId"] = {"not": None}
        
        try:
            # Buscar usuários
            users = await self.db.user.find_many(
                where=where,
                include={
                    "cards": include_cards,
                    "qrcodes": include_cards,
                    "userAccessRules": {
                        "include": {"accessRule": True}
                    } if include_access_rules else False
                },
                order_by={"name": "asc"}
            )
            
            # Processar dados
            processed_users = []
            now = datetime.now()
            
            for user in users:
                # Determinar status
                user_status = self._determine_user_status(user, now)
                
                # Aplicar filtro de status
                if status_filter and status_filter != "all":
                    if user_status != status_filter:
                        continue
                
                user_data = {
                    "id": user.id,
                    "idFaceId": user.idFaceId,
                    "name": user.name,
                    "registration": user.registration,
                    "status": user_status,
                    "hasImage": bool(user.image),
                    "isSynced": bool(user.idFaceId),
                    "createdAt": user.createdAt.isoformat(),
                    "updatedAt": user.updatedAt.isoformat()
                }
                
                # Adicionar período de validade
                if user.beginTime:
                    user_data["beginTime"] = user.beginTime.isoformat()
                if user.endTime:
                    user_data["endTime"] = user.endTime.isoformat()
                
                # Adicionar cartões
                if include_cards:
                    user_data["cards"] = [
                        {"id": c.id, "value": str(c.value)}
                        for c in user.cards
                    ] if user.cards else []
                    user_data["qrcodes"] = [
                        {"id": q.id, "value": q.value}
                        for q in user.qrcodes
                    ] if user.qrcodes else []
                    user_data["totalCards"] = len(user_data["cards"])
                    user_data["totalQRCodes"] = len(user_data["qrcodes"])
                
                # Adicionar regras de acesso
                if include_access_rules and user.userAccessRules:
                    user_data["accessRules"] = [
                        {
                            "id": uar.accessRule.id,
                            "name": uar.accessRule.name,
                            "type": uar.accessRule.type,
                            "priority": uar.accessRule.priority
                        }
                        for uar in user.userAccessRules
                    ]
                    user_data["totalAccessRules"] = len(user_data["accessRules"])
                
                processed_users.append(user_data)
            
            # Calcular estatísticas
            statistics = self._calculate_user_statistics(processed_users)
            
            duration = (datetime.now() - start_time).total_seconds()
            
            report_data = {
                "report_type": "users",
                "generated_at": datetime.now().isoformat(),
                "duration_seconds": duration,
                "filters": {
                    "start_date": start_date.isoformat() if start_date else None,
                    "end_date": end_date.isoformat() if end_date else None,
                    "status_filter": status_filter,
                    "with_image": with_image,
                    "synced_only": synced_only
                },
                "statistics": statistics,
                "users": processed_users
            }
            
            # Formatar saída
            if format_type == "csv":
                return self._format_users_csv(report_data)
            elif format_type == "excel":
                return self._format_users_excel(report_data)
            else:
                return {
                    "success": True,
                    "format": "json",
                    "data": report_data
                }
            
        except Exception as e:
            logger.error(f"Erro ao gerar relatório de usuários: {e}")
            return {
                "success": False,
                "error": str(e)
            }
    
    def _determine_user_status(self, user, now: datetime) -> str:
        """Determina status do usuário"""
        if user.beginTime and user.beginTime > now:
            return "pending"
        
        if user.endTime and user.endTime < now:
            return "expired"
        
        return "active"
    
    def _calculate_user_statistics(self, users: List[Dict]) -> Dict:
        """Calcula estatísticas gerais dos usuários"""
        total = len(users)
        
        if total == 0:
            return {
                "total": 0,
                "by_status": {},
                "with_image": 0,
                "synced": 0,
                "with_cards": 0
            }
        
        status_counts = Counter(u["status"] for u in users)
        
        return {
            "total": total,
            "by_status": dict(status_counts),
            "with_image": sum(1 for u in users if u["hasImage"]),
            "synced": sum(1 for u in users if u["isSynced"]),
            "with_cards": sum(1 for u in users if u.get("totalCards", 0) > 0),
            "percentages": {
                "with_image": round((sum(1 for u in users if u["hasImage"]) / total) * 100, 2),
                "synced": round((sum(1 for u in users if u["isSynced"]) / total) * 100, 2),
                "active": round((status_counts.get("active", 0) / total) * 100, 2)
            }
        }
    
    # ==================== RELATÓRIO DE ACESSOS ====================
    
    async def generate_access_report(
        self,
        start_date: datetime,
        end_date: datetime,
        user_ids: Optional[List[int]] = None,
        portal_ids: Optional[List[int]] = None,
        events: Optional[List[str]] = None,
        group_by: str = "day",  # "day", "hour", "user", "portal", "event"
        include_details: bool = True,
        format_type: str = "json"
    ) -> Dict[str, Any]:
        """
        Gera relatório de acessos (logs)
        
        Args:
            start_date: Data/hora inicial
            end_date: Data/hora final
            user_ids: Filtrar por usuários específicos
            portal_ids: Filtrar por portais específicos
            events: Filtrar por tipos de evento
            group_by: Agrupar por (dia, hora, usuário, portal, evento)
            include_details: Incluir detalhes dos registros
            format_type: Formato de saída
        
        Returns:
            Relatório de acessos
        """
        logger.info(f"Gerando relatório de acessos: {start_date} a {end_date}")
        start_time = datetime.now()
        
        # Construir filtros
        where = {
            "timestamp": {
                "gte": start_date,
                "lte": end_date
            }
        }
        
        if user_ids:
            where["userId"] = {"in": user_ids}
        
        if portal_ids:
            where["portalId"] = {"in": portal_ids}
        
        if events:
            where["event"] = {"in": events}
        
        try:
            # Buscar logs
            logs = await self.db.accesslog.find_many(
                where=where,
                include={
                    "user": True,
                    "portal": True
                },
                order_by={"timestamp": "asc"}
            )
            
            total_logs = len(logs)
            
            if total_logs == 0:
                return {
                    "success": True,
                    "message": "Nenhum registro encontrado no período",
                    "data": {
                        "report_type": "access",
                        "total_logs": 0,
                        "period": {
                            "start": start_date.isoformat(),
                            "end": end_date.isoformat()
                        }
                    }
                }
            
            # Processar logs
            processed_logs = []
            
            for log in logs:
                log_data = {
                    "id": log.id,
                    "timestamp": log.timestamp.isoformat(),
                    "event": log.event,
                    "userId": log.userId,
                    "userName": log.user.name if log.user else "Desconhecido",
                    "portalId": log.portalId,
                    "portalName": log.portal.name if log.portal else "N/A",
                    "cardValue": log.cardValue,
                    "reason": log.reason
                }
                processed_logs.append(log_data)
            
            # Calcular estatísticas
            statistics = self._calculate_access_statistics(logs, group_by)
            
            # Análise temporal
            temporal_analysis = self._analyze_temporal_patterns(logs)
            
            # Top usuários e portais
            top_analysis = self._analyze_top_entities(logs)
            
            duration = (datetime.now() - start_time).total_seconds()
            
            report_data = {
                "report_type": "access",
                "generated_at": datetime.now().isoformat(),
                "duration_seconds": duration,
                "period": {
                    "start": start_date.isoformat(),
                    "end": end_date.isoformat(),
                    "days": (end_date - start_date).days + 1
                },
                "filters": {
                    "user_ids": user_ids,
                    "portal_ids": portal_ids,
                    "events": events
                },
                "statistics": statistics,
                "temporal_analysis": temporal_analysis,
                "top_analysis": top_analysis
            }
            
            # Incluir detalhes se solicitado
            if include_details:
                report_data["logs"] = processed_logs
            
            # Formatar saída
            if format_type == "csv":
                return self._format_access_csv(report_data)
            elif format_type == "excel":
                return self._format_access_excel(report_data)
            else:
                return {
                    "success": True,
                    "format": "json",
                    "data": report_data
                }
            
        except Exception as e:
            logger.error(f"Erro ao gerar relatório de acessos: {e}")
            return {
                "success": False,
                "error": str(e)
            }
    
    def _calculate_access_statistics(self, logs: List, group_by: str) -> Dict:
        """Calcula estatísticas de acessos"""
        total = len(logs)
        
        # Por evento
        events = Counter(log.event for log in logs)
        
        # Por dia da semana
        days_of_week = Counter(log.timestamp.strftime("%A") for log in logs)
        
        stats = {
            "total": total,
            "by_event": {
                event: {
                    "count": count,
                    "percentage": round((count / total) * 100, 2)
                }
                for event, count in events.items()
            },
            "by_day_of_week": dict(days_of_week),
            "success_rate": round(
                (events.get("access_granted", 0) / total * 100), 2
            ) if total > 0 else 0
        }
        
        return stats
    
    def _analyze_temporal_patterns(self, logs: List) -> Dict:
        """Analisa padrões temporais"""
        # Por hora do dia
        by_hour = defaultdict(int)
        for log in logs:
            by_hour[log.timestamp.hour] += 1
        
        # Por dia
        by_date = defaultdict(lambda: {"granted": 0, "denied": 0, "total": 0})
        for log in logs:
            date_key = log.timestamp.strftime("%Y-%m-%d")
            by_date[date_key]["total"] += 1
            if log.event == "access_granted":
                by_date[date_key]["granted"] += 1
            elif log.event == "access_denied":
                by_date[date_key]["denied"] += 1
        
        # Horário de pico
        if by_hour:
            peak_hour = max(by_hour.items(), key=lambda x: x[1])
        else:
            peak_hour = (0, 0)
        
        return {
            "by_hour": dict(sorted(by_hour.items())),
            "by_date": dict(sorted(by_date.items())),
            "peak_hour": {
                "hour": peak_hour[0],
                "count": peak_hour[1]
            }
        }
    
    def _analyze_top_entities(self, logs: List, limit: int = 10) -> Dict:
        """Analisa top usuários e portais"""
        # Top usuários
        user_counts = defaultdict(lambda: {"granted": 0, "denied": 0, "name": None})
        for log in logs:
            if log.userId:
                if log.event == "access_granted":
                    user_counts[log.userId]["granted"] += 1
                elif log.event == "access_denied":
                    user_counts[log.userId]["denied"] += 1
                
                if log.user:
                    user_counts[log.userId]["name"] = log.user.name
        
        top_users = sorted(
            [
                {
                    "userId": uid,
                    "userName": data["name"] or f"User {uid}",
                    "totalAccess": data["granted"] + data["denied"],
                    "granted": data["granted"],
                    "denied": data["denied"]
                }
                for uid, data in user_counts.items()
            ],
            key=lambda x: x["totalAccess"],
            reverse=True
        )[:limit]
        
        # Top portais
        portal_counts = defaultdict(lambda: {"granted": 0, "denied": 0, "name": None})
        for log in logs:
            if log.portalId:
                if log.event == "access_granted":
                    portal_counts[log.portalId]["granted"] += 1
                elif log.event == "access_denied":
                    portal_counts[log.portalId]["denied"] += 1
                
                if log.portal:
                    portal_counts[log.portalId]["name"] = log.portal.name
        
        top_portals = sorted(
            [
                {
                    "portalId": pid,
                    "portalName": data["name"] or f"Portal {pid}",
                    "totalAccess": data["granted"] + data["denied"],
                    "granted": data["granted"],
                    "denied": data["denied"]
                }
                for pid, data in portal_counts.items()
            ],
            key=lambda x: x["totalAccess"],
            reverse=True
        )[:limit]
        
        return {
            "top_users": top_users,
            "top_portals": top_portals
        }
    
    # ==================== FORMATAÇÃO CSV ====================
    
    def _format_users_csv(self, report_data: Dict) -> Dict[str, Any]:
        """Formata relatório de usuários em CSV"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([
            "ID",
            "Nome",
            "Matrícula",
            "Status",
            "Tem Imagem",
            "Sincronizado",
            "Total Cartões",
            "Total QR Codes",
            "Total Regras",
            "Criado Em",
            "Atualizado Em"
        ])
        
        # Dados
        for user in report_data["users"]:
            writer.writerow([
                user["id"],
                user["name"],
                user.get("registration", ""),
                user["status"],
                "Sim" if user["hasImage"] else "Não",
                "Sim" if user["isSynced"] else "Não",
                user.get("totalCards", 0),
                user.get("totalQRCodes", 0),
                user.get("totalAccessRules", 0),
                user["createdAt"],
                user["updatedAt"]
            ])
        
        csv_content = output.getvalue()
        output.close()
        
        return {
            "success": True,
            "format": "csv",
            "content": csv_content,
            "filename": f"relatorio_usuarios_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "statistics": report_data["statistics"]
        }
    
    def _format_access_csv(self, report_data: Dict) -> Dict[str, Any]:
        """Formata relatório de acessos em CSV"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([
            "ID",
            "Data/Hora",
            "Evento",
            "Usuário ID",
            "Usuário Nome",
            "Portal ID",
            "Portal Nome",
            "Cartão",
            "Motivo"
        ])
        
        # Dados
        for log in report_data.get("logs", []):
            writer.writerow([
                log["id"],
                log["timestamp"],
                log["event"],
                log["userId"] or "",
                log["userName"],
                log["portalId"] or "",
                log["portalName"],
                log["cardValue"] or "",
                log["reason"] or ""
            ])
        
        csv_content = output.getvalue()
        output.close()
        
        return {
            "success": True,
            "format": "csv",
            "content": csv_content,
            "filename": f"relatorio_acessos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "statistics": report_data["statistics"]
        }
    
    def _format_users_excel(self, report_data: Dict) -> Dict[str, Any]:
        """Formata relatório de usuários em Excel (requer openpyxl)"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Usuários"
            
            # Estilo do cabeçalho
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            # Cabeçalho
            headers = ["ID", "Nome", "Matrícula", "Status", "Tem Imagem", "Sincronizado", 
                       "Cartões", "QR Codes", "Regras", "Criado Em"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Dados
            for row_idx, user in enumerate(report_data["users"], 2):
                ws.cell(row=row_idx, column=1, value=user["id"])
                ws.cell(row=row_idx, column=2, value=user["name"])
                ws.cell(row=row_idx, column=3, value=user.get("registration", ""))
                ws.cell(row=row_idx, column=4, value=user["status"])
                ws.cell(row=row_idx, column=5, value="Sim" if user["hasImage"] else "Não")
                ws.cell(row=row_idx, column=6, value="Sim" if user["isSynced"] else "Não")
                ws.cell(row=row_idx, column=7, value=user.get("totalCards", 0))
                ws.cell(row=row_idx, column=8, value=user.get("totalQRCodes", 0))
                ws.cell(row=row_idx, column=9, value=user.get("totalAccessRules", 0))
                ws.cell(row=row_idx, column=10, value=user["createdAt"])
            
            # Ajustar largura das colunas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            # Salvar em buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            excel_content = buffer.getvalue()
            buffer.close()
            
            return {
                "success": True,
                "format": "excel",
                "content": excel_content,
                "filename": f"relatorio_usuarios_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "statistics": report_data["statistics"]
            }
            
        except ImportError:
            logger.warning("openpyxl não instalado. Retornando CSV.")
            return self._format_users_csv(report_data)
    
    def _format_access_excel(self, report_data: Dict) -> Dict[str, Any]:
        """Formata relatório de acessos em Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            
            # Aba 1: Logs
            ws_logs = wb.active
            ws_logs.title = "Logs de Acesso"
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            headers = ["ID", "Data/Hora", "Evento", "Usuário", "Portal", "Cartão", "Motivo"]
            
            for col, header in enumerate(headers, 1):
                cell = ws_logs.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            for row_idx, log in enumerate(report_data.get("logs", []), 2):
                ws_logs.cell(row=row_idx, column=1, value=log["id"])
                ws_logs.cell(row=row_idx, column=2, value=log["timestamp"])
                ws_logs.cell(row=row_idx, column=3, value=log["event"])
                ws_logs.cell(row=row_idx, column=4, value=log["userName"])
                ws_logs.cell(row=row_idx, column=5, value=log["portalName"])
                ws_logs.cell(row=row_idx, column=6, value=log["cardValue"] or "")
                ws_logs.cell(row=row_idx, column=7, value=log["reason"] or "")
            
            # Aba 2: Estatísticas
            ws_stats = wb.create_sheet("Estatísticas")
            
            stats = report_data["statistics"]
            ws_stats.cell(row=1, column=1, value="Estatística").font = Font(bold=True)
            ws_stats.cell(row=1, column=2, value="Valor").font = Font(bold=True)
            
            ws_stats.cell(row=2, column=1, value="Total de Acessos")
            ws_stats.cell(row=2, column=2, value=stats["total"])
            
            ws_stats.cell(row=3, column=1, value="Taxa de Sucesso")
            ws_stats.cell(row=3, column=2, value=f"{stats['success_rate']}%")
            
            # Salvar
            buffer = io.BytesIO()
            wb.save(buffer)
            excel_content = buffer.getvalue()
            buffer.close()
            
            return {
                "success": True,
                "format": "excel",
                "content": excel_content,
                "filename": f"relatorio_acessos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "statistics": report_data["statistics"]
            }
            
        except ImportError:
            return self._format_access_csv(report_data)
    
    # ==================== RELATÓRIOS RÁPIDOS ====================
    
    async def quick_daily_report(self, date: Optional[datetime] = None) -> Dict:
        """Relatório rápido do dia"""
        if not date:
            date = datetime.now()
        
        start_of_day = date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_of_day = date.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        return await self.generate_access_report(
            start_date=start_of_day,
            end_date=end_of_day,
            include_details=False,
            format_type="json"
        )
    
    async def quick_weekly_report(self) -> Dict:
        """Relatório rápido da semana"""
        today = datetime.now()
        start_of_week = today - timedelta(days=today.weekday())
        start_of_week = start_of_week.replace(hour=0, minute=0, second=0, microsecond=0)
        
        return await self.generate_access_report(
            start_date=start_of_week,
            end_date=today,
            include_details=False,
            format_type="json"
        )
    
    async def quick_monthly_report(self) -> Dict:
        """Relatório rápido do mês"""
        today = datetime.now()
        start_of_month = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        return await self.generate_access_report(
            start_date=start_of_month,
            end_date=today,
            include_details=False,
            format_type="json"
        )