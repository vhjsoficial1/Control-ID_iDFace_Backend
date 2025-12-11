import asyncio
import openpyxl
import logging
import os
import sys

# Adicionar o diretório raiz do projeto ao sys.path
# Isso permite que o script encontre os módulos da aplicação (app, etc.)
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.database import db, connect_db, disconnect_db
from app.services.user_service import UserService

# Configuração do logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# --- CONFIGURAÇÃO ---
# Caminho para o arquivo Excel contendo os usuários
# A primeira coluna deve ser o NOME e a segunda a MATRÍCULA
EXCEL_FILE_PATH = r"C:\Users\vitor.hsantos\OneDrive - Sistema FIEB\Documentos\Vitor\Automação\MaxPlast\empregados import.xlsx"


async def main():
    """
    Script principal para ler o arquivo Excel e importar os usuários para o banco de dados.
    """
    logging.info("Iniciando processo de importação de usuários...")

    if not os.path.exists(EXCEL_FILE_PATH):
        logging.error(f"Arquivo não encontrado: {EXCEL_FILE_PATH}")
        return

    try:
        # Conectar ao banco de dados
        await connect_db()
        user_service = UserService(db)

        # Carregar a planilha Excel
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
        sheet = workbook.active
        
        logging.info(f"Arquivo Excel '{os.path.basename(EXCEL_FILE_PATH)}' carregado com sucesso.")
        
        # Iterar pelas linhas da planilha (começando da segunda linha para pular o cabeçalho)
        # Assumindo: Coluna A = Nome, Coluna B = Matrícula
        total_rows = sheet.max_row
        imported_count = 0
        failed_count = 0

        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            name = row[0]
            registration = row[1]

            # Validar dados da linha
            if not name or not isinstance(name, str) or len(name.strip()) == 0:
                logging.warning(f"Linha {i}: Nome inválido ou ausente. Pulando.")
                failed_count += 1
                continue

            # A matrícula pode ser um número, então convertemos para string
            if registration is not None:
                registration = str(registration).strip()
            else:
                registration = None # Garante que seja nulo se a célula estiver vazia

            logging.info(f"Processando Linha {i}: Nome='{name}', Matrícula='{registration}'")

            try:
                # Verificar se um usuário com a mesma matrícula já existe
                if registration:
                    existing_user_result = await user_service.search_users(registration=registration)
                    if existing_user_result.get("total", 0) > 0:
                        logging.warning(f"Linha {i}: Usuário com matrícula '{registration}' já existe. Pulando.")
                        failed_count += 1
                        continue

                # Criar o usuário
                result = await user_service.create_user(
                    name=name.strip(),
                    registration=registration
                )

                if result["success"]:
                    logging.info(f"SUCESSO: Usuário '{name}' criado (ID: {result['user'].id}).")
                    imported_count += 1
                else:
                    logging.error(f"FALHA ao criar usuário '{name}': {result.get('errors', 'Erro desconhecido')}")
                    failed_count += 1

            except Exception as e:
                logging.error(f"FALHA INESPERADA na linha {i} para o usuário '{name}': {e}")
                failed_count += 1

        logging.info("--- Resumo da Importação ---")
        logging.info(f"Total de linhas processadas: {total_rows - 1}")
        logging.info(f"Usuários criados com sucesso: {imported_count}")
        logging.info(f"Registros com falha ou já existentes: {failed_count}")
        logging.info("-----------------------------")


    except Exception as e:
        logging.error(f"Ocorreu um erro crítico durante o processo: {e}")
    finally:
        # Desconectar do banco de dados
        await disconnect_db()
        logging.info("Processo de importação finalizado.")


if __name__ == "__main__":
    # Adiciona um loop de eventos asyncio se não houver um
    try:
        asyncio.run(main())
    except RuntimeError as e:
        if "cannot run loop while another loop is running" in str(e):
            # Se já houver um loop, apenas agendamos a execução
            # Isso pode ser útil em alguns ambientes
            loop = asyncio.get_event_loop()
            loop.create_task(main())
        else:
            raise
